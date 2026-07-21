#!/usr/bin/env python3
"""v21: Fair few-shot geochem extraction — same inputs as Ours pipeline.

Differences from v17/v20:
- Reads PDF text AND supplementary spreadsheet(s) — same input bundle
  the Ours pipeline gets.
- Single LLM call per paper with 32k input budget (PDF + supp serialised
  as compact CSV-ish table strings).
- 3-shot ICL using the geochem 210-column schema as the target format
  so output structure matches what the strict 4-tier evaluator expects.
- 16k max_tokens output to allow ~50-150 sample rows per call.

This is the FAIR few-shot baseline: same inputs as Ours, no architectural
machinery (no ontology, no consensus, no self-correction), single LLM
call. Scored with the strict geochem 4-tier evaluator.
"""
from __future__ import annotations
import os, sys, json, time, traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT.parents[0]))  # src/ -> articleminer

env = ROOT / ".env"
if env.exists():
    for line in env.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ[k] = v.strip().strip('"').strip("'")

from articleminer.pdf_reader import extract_pdf, get_paper_text_for_llm
from pipeline_adapter import make_client
import openpyxl

GT_DIR       = ROOT.parents[1] / "data" / "geochem28" / "ground_truth"
GEOCHEM_DATA = ROOT.parents[1] / "data" / "geochem28" / "pdfs"
SPREAD_DIR   = GEOCHEM_DATA / "Spreadsheets"
RESULTS_DIR  = ROOT.parents[1] / "results"

PROMPT = """You are extracting geochemical analyses for a knowledge graph.

You will receive (a) PDF text and (b) supplementary spreadsheet(s) for ONE paper.
Your job: emit ONE JSON object per analysed sample, in the schema below.

OUTPUT SCHEMA (one JSON object per line, NO commentary):
{"sample_id":"<exact ID>","mineral":"<host mineral, lowercase>",
 "method":"<analytical method, e.g. LA-ICPMS>","deposit":"<deposit name>",
 "deposit_type":"<deposit class, e.g. MVT zinc-lead>","elements":{"<el>":<ppm>,...}}

Rules:
- Use ppm for trace elements; convert wt% values multiply by 10000.
- Use the SAMPLE ID exactly as written (do not invent or shorten).
- For below-detection-limit values, use the negative of the limit (e.g. -0.5).
- Skip detection-limit rows, summary rows, reference standards (NIST 610 etc).
- One JSON line per sample. Output every sample present in the supplementary
  spreadsheet.

EXAMPLES (truncated):
{"sample_id":"WG-1","mineral":"sphalerite","method":"LA-ICPMS","deposit":"Daliangzi","deposit_type":"MVT zinc-lead","elements":{"Cu":12.3,"Zn":650000,"Pb":4500,"Ag":-0.1}}
{"sample_id":"WG-2","mineral":"sphalerite","method":"LA-ICPMS","deposit":"Daliangzi","deposit_type":"MVT zinc-lead","elements":{"Cu":8.2,"Zn":640000,"Pb":3900,"Ag":0.8}}"""


def list_gt_papers():
    return sorted([p for p in GT_DIR.glob("*.xlsx") if not p.name.startswith("_")])


def find_pdf(paper_id: str):
    for ext in (".pdf", ".PDF"):
        f = GEOCHEM_DATA / f"{paper_id}{ext}"
        if f.exists(): return f
    matches = list(GEOCHEM_DATA.glob(f"{paper_id}*.pdf"))
    if matches: return matches[0]
    stem = paper_id.split("_et_al")[0]
    matches = list(GEOCHEM_DATA.glob(f"*{stem}*.pdf"))
    if matches: return matches[0]
    return None


def find_supp(paper_id: str) -> list[Path]:
    """Return list of supplementary files matching the paper id."""
    if not SPREAD_DIR.exists(): return []
    out = []
    stem_keys = [paper_id]
    parts = paper_id.split("_et_al")
    if parts:
        stem_keys.append(parts[0])
    all_files = list(SPREAD_DIR.glob("*.xlsx")) + list(SPREAD_DIR.glob("*.csv"))
    # Try multiple match keys: full id, stem, year+author
    keys = [paper_id.lower()]
    if "_et_al" in paper_id:
        base = paper_id.split("_et_al")[0].lower()
        keys.append(base)
        if "_et_al_" in paper_id:
            year = paper_id.split("_et_al_")[-1].split("_")[0]
            keys.append(f"{year}_{base}")
            keys.append(f"{base}_{year}")
    for f in all_files:
        nm = f.name.lower()
        if any(k in nm for k in keys):
            out.append(f)
    return out


def serialize_xlsx(f: Path, max_rows: int = 80, max_chars: int = 12000) -> str:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        return f"[xlsx read err: {e}]"
    chunks = [f"=== FILE: {f.name} ==="]
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))[:max_rows]
        if not rows: continue
        chunks.append(f"--- sheet: {sn} ({len(rows)} rows) ---")
        for r in rows:
            line = ",".join("" if c is None else str(c)[:25] for c in r[:50])
            chunks.append(line)
        if len("\n".join(chunks)) > max_chars: break
    wb.close()
    return "\n".join(chunks)[:max_chars]


def serialize_csv(f: Path, max_chars: int = 12000) -> str:
    try:
        text = f.read_text(errors="ignore")[:max_chars]
        return f"=== FILE: {f.name} (csv) ===\n{text}"
    except Exception as e:
        return f"[csv read err: {e}]"


def parse_jsonl(raw: str) -> list[dict]:
    rows = []
    for line in raw.splitlines():
        line = line.strip().rstrip(",")
        if not line.startswith("{"): continue
        try:
            obj = json.loads(line)
            if "sample_id" in obj:
                rows.append(obj)
        except Exception:
            continue
    return rows


def run_one_paper(paper_id: str, model: str, client) -> dict:
    pdf = find_pdf(paper_id)
    if not pdf:
        return {"paper": paper_id, "model": model, "error": "no_pdf"}
    try:
        content = extract_pdf(str(pdf))
        pdf_text = get_paper_text_for_llm(content)[:25000]
    except Exception as e:
        return {"paper": paper_id, "model": model, "error": f"pdf_read: {e}"}

    supp_files = find_supp(paper_id)
    supp_serialized = []
    budget = 25000
    for sf in supp_files[:4]:
        if sf.suffix.lower() == ".csv":
            txt = serialize_csv(sf, max_chars=budget // max(1, len(supp_files)))
        else:
            txt = serialize_xlsx(sf, max_chars=budget // max(1, len(supp_files)))
        supp_serialized.append(txt)
    supp_block = "\n\n".join(supp_serialized) if supp_serialized else "[no supplementary files]"

    user = f"PDF TEXT:\n{pdf_text}\n\n--- SUPPLEMENTARY DATA ---\n{supp_block}\n\nExtract every sample analysis as JSON lines:"
    try:
        raw = client.complete(PROMPT, user, max_tokens=16384)
    except Exception as e:
        return {"paper": paper_id, "model": model, "error": f"llm: {e}"}

    rows = parse_jsonl(raw)
    return {"paper": paper_id, "model": model, "n_rows": len(rows),
            "rows": rows, "n_supp": len(supp_files),
            "supp_files": [s.name for s in supp_files],
            "raw_preview": raw[:400]}


def main(model_filter=None):
    papers = list_gt_papers()
    models = [
        "claude-sonnet-4-6",
        "claude-opus-4-6",
        "claude-haiku-4-5-20251001",
        "gpt-4o",
        "gemini-2.5-flash",
        "qwen25-7b",
        "mistral-7b-v02",
    ]
    if model_filter:
        models = [m for m in models if model_filter in m]
    for model in models:
        out_dir = RESULTS_DIR / f"geochem_fewshot_pdfsupp_{model.replace('/', '_')}_v21"
        out_dir.mkdir(parents=True, exist_ok=True)
        print(f"\n=== {model} → {out_dir.name} ===", flush=True)
        try:
            client = make_client(model)
        except Exception as e:
            print(f"  SKIP {model}: client init failed ({e})")
            continue
        t0 = time.time()
        for i, gt in enumerate(papers, 1):
            paper_id = gt.stem
            f_out = out_dir / f"{paper_id}.json"
            if f_out.exists():
                continue
            t1 = time.time()
            res = run_one_paper(paper_id, model, client)
            res["elapsed_seconds"] = round(time.time() - t1, 1)
            json.dump(res, open(f_out, "w"), indent=2, default=str)
            print(f"  [{i}/{len(papers)}] {paper_id}: {res.get('n_rows','?')} rows, "
                  f"supp={res.get('n_supp',0)} ({res['elapsed_seconds']:.0f}s)"
                  + (f"  ERR={res['error']}" if "error" in res else ""), flush=True)
        print(f"  done in {(time.time()-t0)/60:.1f} min", flush=True)


if __name__ == "__main__":
    flt = sys.argv[1] if len(sys.argv) > 1 else None
    main(flt)
